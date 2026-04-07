#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Download Binance coin-margined perpetual funding rates, aggregate by day and month,
and export symbol-level daily Excel files plus summary workbooks.
"""

from __future__ import annotations

import argparse
from copy import copy
import math
import tempfile
import shutil
import time
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Dict, List, Set, Tuple

from binance.client import Client
from binance.exceptions import BinanceAPIException, BinanceRequestException
import pandas as pd
from requests.exceptions import RequestException
from openpyxl.chart import LineChart, Reference
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlite_store import (
    build_weekly_symbol_metrics,
    build_monthly_symbol_metrics,
    create_collector_run,
    finalize_collector_run,
    initialize_database,
    load_daily_funding_metrics,
    persist_daily_funding_metrics,
    persist_daily_volume_metrics,
    persist_funding_quality_audit,
    persist_volume_quality_audit,
    persist_market_snapshots,
    persist_monthly_funding_metrics,
    persist_weekly_funding_metrics,
    persist_raw_funding_rates,
    sqlite_connection,
    upsert_symbols,
)

FULL_LOOKBACK_YEARS = 3
DEFAULT_VOLUME_LOOKBACK_EXTRA_DAYS = 7
API_BATCH_LIMIT = 1000
REQUEST_SLEEP_SECONDS = 0.2
MAX_RETRIES = 5

OUTPUT_ROOT = Path(__file__).resolve().parent / "coin_funding_rate_outputs"
DAILY_DIR = OUTPUT_ROOT / "daily"

SUMMARY_BASENAME = "\u8d39\u7387\u7edf\u8ba1\u8868"
SUMMARY_TIME_FORMAT = "%Y%m%d%H%M"

GROUP_TIMEZONE = "Asia/Shanghai"
AVG_VOLUME_WINDOW_DAYS = 30
VOLUME_KLINE_BATCH_LIMIT = 1500
VOLUME_KLINE_MAX_WINDOW_DAYS = 200
LOW_VOLUME_THRESHOLD = 10_000_000
DAILY_CHART_DAYS = 30

MONTHLY_SHEET_NAME = "MonthlySummary"
RANKING_SHEET_NAME = "TopRanking"
OVERVIEW_SHEET_NAME = "Overview"
VOLUME_SHEET_NAME = "30\u5929\u65e5\u5e73\u5747\u6210\u4ea4\u91cf"
LOW_VOLUME_MARK = "*"

HIGHLIGHT_FILL = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
LOW_VOLUME_FILL = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
POSITIVE_FONT = Font(color="FF1B5E20")
NEGATIVE_FONT = Font(color="FFB71C1C")
FOCUS_BASKET = ["BTCUSD_PERP", "ETHUSD_PERP", "BNBUSD_PERP", "SOLUSD_PERP", "XRPUSD_PERP"]


class DataFetchIncompleteError(RuntimeError):
    pass


def simplify_symbol(symbol: str) -> str:
    base = symbol.replace("_PERP", "")
    if base.endswith("USD"):
        base = base[:-3]
    return base


def make_safe_sheet_name(base: str, used_names: Set[str]) -> str:
    invalid_chars = set(r'[]:*?/\\')
    cleaned = "".join(ch for ch in base if ch not in invalid_chars).strip()
    if not cleaned:
        cleaned = "Sheet"
    cleaned = cleaned[:31]
    candidate = cleaned
    idx = 1
    while candidate in used_names:
        suffix = f"_{idx}"
        candidate = f"{cleaned[:31 - len(suffix)]}{suffix}"
        idx += 1
    used_names.add(candidate)
    return candidate


def get_coin_perpetual_symbols(client: Client) -> Tuple[List[str], Dict[str, float]]:
    try:
        info = client.futures_coin_exchange_info()
    except Exception as exc:  # noqa: BLE001
        raise RuntimeError(f"\u65e0\u6cd5\u83b7\u53d6\u4ea4\u6613\u5bf9\u5217\u8868: {exc}") from exc

    symbols: List[str] = []
    contract_sizes: Dict[str, float] = {}
    for symbol_info in info.get("symbols", []):
        if symbol_info.get("contractType") == "PERPETUAL" and symbol_info.get("contractStatus") == "TRADING":
            symbol = symbol_info["symbol"]
            symbols.append(symbol)
            try:
                contract_sizes[symbol] = float(symbol_info.get("contractSize", 1.0))
            except (TypeError, ValueError):
                contract_sizes[symbol] = 1.0

    unique_symbols = sorted(set(symbols))
    for sym in unique_symbols:
        contract_sizes.setdefault(sym, 1.0)
    return unique_symbols, contract_sizes


def fetch_symbol_funding_history(client: Client, symbol: str, start_ts: int) -> pd.DataFrame:
    now_ms = int(datetime.now(timezone.utc).timestamp() * 1000)
    cursor = start_ts
    rows: List[dict] = []

    while cursor < now_ms:
        retries = 0
        while retries < MAX_RETRIES:
            try:
                batch = client.futures_coin_funding_rate(symbol=symbol, startTime=cursor, limit=API_BATCH_LIMIT)
                break
            except (BinanceAPIException, BinanceRequestException, RequestException) as exc:
                retries += 1
                wait = REQUEST_SLEEP_SECONDS * (2**retries)
                print(f"[{symbol}] API\u9519\u8bef\uff0c{wait:.1f}s\u540e\u91cd\u8bd5 ({retries}/{MAX_RETRIES}): {exc}")
                time.sleep(wait)
        else:
            raise DataFetchIncompleteError(f"[{symbol}] \u591a\u6b21\u91cd\u8bd5\u5931\u8d25\uff0c\u8be5\u4ea4\u6613\u5bf9\u7684\u8d44\u91d1\u8d39\u7387\u6570\u636e\u4e0d\u5b8c\u6574\u3002")

        if not batch:
            break

        rows.extend(batch)
        cursor = int(batch[-1]["fundingTime"]) + 1
        if len(batch) < API_BATCH_LIMIT and cursor >= now_ms:
            break
        time.sleep(REQUEST_SLEEP_SECONDS)

    if not rows:
        return pd.DataFrame(columns=["fundingTime", "fundingRate"])

    df = pd.DataFrame(rows)
    raw_row_count = len(df)
    df["fundingTime"] = pd.to_datetime(df["fundingTime"], unit="ms", utc=True)
    df["fundingRate"] = pd.to_numeric(df["fundingRate"], errors="coerce").fillna(0.0)
    df.drop_duplicates(subset="fundingTime", inplace=True)
    df.attrs["raw_row_count"] = raw_row_count
    return df


def build_funding_quality_audit(symbol: str, funding_df: pd.DataFrame, daily_df: pd.DataFrame, raw_row_count: int) -> Dict[str, object]:
    duplicate_event_count = max(raw_row_count - len(funding_df), 0)
    if funding_df.empty:
        return {
            "symbol": symbol,
            "raw_event_count": raw_row_count,
            "duplicate_event_count": duplicate_event_count,
            "first_funding_time": None,
            "last_funding_time": None,
            "inferred_interval_hours": 0.0,
            "gap_count": 0,
            "max_gap_hours": 0.0,
            "day_count": 0,
            "days_with_zero_events": 0,
            "min_events_per_day": 0,
            "max_events_per_day": 0,
            "completeness_score": 0.0,
            "status": "empty",
            "notes": "no funding rows returned",
        }

    times = funding_df["fundingTime"].sort_values()
    first_funding_time = pd.Timestamp(times.iloc[0]).isoformat()
    last_funding_time = pd.Timestamp(times.iloc[-1]).isoformat()
    diffs = times.diff().dropna().dt.total_seconds().div(3600)
    inferred_interval_hours = float(diffs.median()) if not diffs.empty else 0.0
    gap_threshold = inferred_interval_hours * 1.5 if inferred_interval_hours > 0 else 0.0
    large_gaps = diffs[diffs > gap_threshold] if gap_threshold > 0 else diffs.iloc[0:0]
    gap_count = int(len(large_gaps))
    max_gap_hours = float(large_gaps.max()) if gap_count else 0.0

    day_count = int(len(daily_df))
    days_with_zero_events = int((daily_df["funding_event_count"] == 0).sum()) if not daily_df.empty else 0
    min_events_per_day = int(daily_df["funding_event_count"].min()) if not daily_df.empty else 0
    max_events_per_day = int(daily_df["funding_event_count"].max()) if not daily_df.empty else 0

    completeness_score = 100.0
    completeness_score -= min(duplicate_event_count * 0.2, 10.0)
    completeness_score -= min(gap_count * 8.0, 40.0)
    completeness_score -= min(days_with_zero_events * 4.0, 20.0)
    completeness_score = max(completeness_score, 0.0)

    status = "ok"
    notes: List[str] = []
    if duplicate_event_count:
        notes.append(f"duplicates={duplicate_event_count}")
    if gap_count:
        notes.append(f"gaps={gap_count}")
        status = "warning"
    if days_with_zero_events:
        notes.append(f"zero_event_days={days_with_zero_events}")
        status = "warning"
    if inferred_interval_hours <= 0:
        status = "warning"
        notes.append("interval_unknown")

    return {
        "symbol": symbol,
        "raw_event_count": raw_row_count,
        "duplicate_event_count": duplicate_event_count,
        "first_funding_time": first_funding_time,
        "last_funding_time": last_funding_time,
        "inferred_interval_hours": inferred_interval_hours,
        "gap_count": gap_count,
        "max_gap_hours": max_gap_hours,
        "day_count": day_count,
        "days_with_zero_events": days_with_zero_events,
        "min_events_per_day": min_events_per_day,
        "max_events_per_day": max_events_per_day,
        "completeness_score": round(completeness_score, 2),
        "status": status,
        "notes": ", ".join(notes) if notes else "ok",
    }


def build_volume_quality_audit(symbol: str, volume_df: pd.DataFrame, expected_window_days: int) -> Dict[str, object]:
    if volume_df.empty:
        return {
            "symbol": symbol,
            "source_type": "coinm_1d_kline",
            "kline_row_count": 0,
            "first_metric_date": None,
            "last_metric_date": None,
            "day_count": 0,
            "gap_count": 0,
            "max_gap_days": 0,
            "avg_usd_volume": 0.0,
            "min_usd_volume": 0.0,
            "max_usd_volume": 0.0,
            "completeness_score": 0.0,
            "status": "empty",
            "notes": "no kline volume rows returned",
        }

    sorted_volume = volume_df.sort_values("date").drop_duplicates(subset="date", keep="last")
    dates = pd.to_datetime(sorted_volume["date"])
    diffs = dates.diff().dropna().dt.days
    gap_sizes = diffs[diffs > 1]
    gap_count = int(len(gap_sizes))
    max_gap_days = int(gap_sizes.max()) if gap_count else 0
    avg_usd_volume = float(sorted_volume["usd_volume"].mean())
    min_usd_volume = float(sorted_volume["usd_volume"].min())
    max_usd_volume = float(sorted_volume["usd_volume"].max())
    completeness_score = 100.0 - min(gap_count * 5.0, 40.0)
    notes = [
        "usd_volume=contract_volume*contract_size",
        "source=futures_coin_klines",
        "official_kline_fields:v=contract_volume,q=base_asset_volume",
    ]
    status = "ok"
    if gap_count:
        status = "warning"
        notes.append(f"gaps={gap_count}")
    if len(sorted_volume) < expected_window_days:
        notes.append("short_window_or_new_listing")

    return {
        "symbol": symbol,
        "source_type": "coinm_1d_kline",
        "kline_row_count": int(len(sorted_volume)),
        "first_metric_date": pd.Timestamp(dates.iloc[0]).strftime("%Y-%m-%d"),
        "last_metric_date": pd.Timestamp(dates.iloc[-1]).strftime("%Y-%m-%d"),
        "day_count": int(len(sorted_volume)),
        "gap_count": gap_count,
        "max_gap_days": max_gap_days,
        "avg_usd_volume": avg_usd_volume,
        "min_usd_volume": min_usd_volume,
        "max_usd_volume": max_usd_volume,
        "completeness_score": round(max(completeness_score, 0.0), 2),
        "status": status,
        "notes": ", ".join(notes),
    }


def build_failed_volume_quality_audit(symbol: str, note: str) -> Dict[str, object]:
    return {
        "symbol": symbol,
        "source_type": "coinm_1d_kline",
        "kline_row_count": 0,
        "first_metric_date": None,
        "last_metric_date": None,
        "day_count": 0,
        "gap_count": 0,
        "max_gap_days": 0,
        "avg_usd_volume": 0.0,
        "min_usd_volume": 0.0,
        "max_usd_volume": 0.0,
        "completeness_score": 0.0,
        "status": "failed",
        "notes": note,
    }


def compute_daily_funding(df: pd.DataFrame, group_timezone: str = GROUP_TIMEZONE) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame(columns=["date", "daily_funding_rate", "funding_event_count"])

    daily = (
        df.assign(funding_event_count=1)
        .set_index("fundingTime")
        .tz_convert(group_timezone)
        .resample("D")
        .agg({"fundingRate": "sum", "funding_event_count": "sum"})
        .reset_index()
    )
    daily["date"] = daily["fundingTime"].dt.tz_localize(None)
    daily.rename(columns={"fundingRate": "daily_funding_rate"}, inplace=True)
    return daily[["date", "daily_funding_rate", "funding_event_count"]]


def compute_monthly_summary(
    symbol_daily: Dict[str, pd.DataFrame], periods: int = 24, group_timezone: str = GROUP_TIMEZONE
) -> pd.DataFrame:
    monthly_frames = []
    for symbol, daily in symbol_daily.items():
        if daily.empty:
            continue
        temp = daily.copy()
        temp["date"] = pd.to_datetime(temp["date"]).dt.tz_localize(None)
        monthly = temp.set_index("date")["daily_funding_rate"].resample("ME").sum().rename(symbol)
        monthly.index = monthly.index.to_period("M")
        monthly_frames.append(monthly)

    if not monthly_frames:
        return pd.DataFrame()

    monthly_summary = pd.concat(monthly_frames, axis=1).fillna(0.0)
    monthly_summary = monthly_summary.groupby(level=0).sum()
    current_period = pd.Timestamp.now(tz=group_timezone).tz_localize(None).to_period("M")
    period_index = pd.period_range(end=current_period, periods=periods, freq="M")
    monthly_summary = monthly_summary.reindex(period_index, fill_value=0.0)
    monthly_summary = monthly_summary.rename(columns={col: simplify_symbol(col) for col in monthly_summary.columns})
    monthly_summary = monthly_summary.T.groupby(level=0).sum().T
    monthly_summary = monthly_summary[monthly_summary.sum().sort_values(ascending=False).index]
    monthly_summary.index.name = "Month"
    return monthly_summary


def get_window_totals(ordered_monthly: pd.DataFrame, end_offset: int, months: int) -> pd.Series:
    if ordered_monthly.empty or months <= 0:
        return pd.Series(dtype=float)
    period_index = ordered_monthly.index
    if len(period_index) == 0:
        return pd.Series(dtype=float)

    current_period = period_index[-1]
    end_period = current_period - end_offset
    start_period = end_period - (months - 1)
    window = ordered_monthly.loc[(period_index >= start_period) & (period_index <= end_period)]
    if window.empty:
        return pd.Series(dtype=float)

    totals = window.sum()
    totals.index = [simplify_symbol(symbol) for symbol in totals.index]
    return totals.groupby(level=0).sum().sort_values(ascending=False)


def format_top_symbols(totals: pd.Series, top_n: int = 3) -> str:
    if totals.empty:
        return "-"
    return ", ".join([f"{symbol} ({value * 100:.2f}%)" for symbol, value in totals.head(top_n).items()])


def build_recent_top_table(monthly_summary: pd.DataFrame) -> pd.DataFrame:
    if monthly_summary.empty:
        return pd.DataFrame()

    ordered_monthly = monthly_summary.sort_index()
    top_n = 8
    windows = [
        ("\u672c\u6708", 0, 1),
        ("\u4e0a\u4e2a\u6708", 1, 1),
        ("\u4e0a\u4e09\u4e2a\u6708", 1, 3),
        ("\u4e0a6\u4e2a\u6708", 1, 6),
        ("\u4e0a12\u4e2a\u6708", 1, 12),
    ]

    records = []
    for label, end_offset, months in windows:
        totals = get_window_totals(ordered_monthly, end_offset=end_offset, months=months)
        top_entries = [f"{symbol} ({value * 100:.2f}%)" for symbol, value in totals.head(top_n).items()]
        while len(top_entries) < top_n:
            top_entries.append("-")
        records.append((label, top_entries))

    return pd.DataFrame({f"Rank{i + 1}": [record[1][i] for record in records] for i in range(top_n)}, index=[r[0] for r in records])


def build_overview_table(monthly_summary: pd.DataFrame, low_volume_symbols: Set[str]) -> pd.DataFrame:
    if monthly_summary.empty:
        return pd.DataFrame(columns=["\u6307\u6807", "\u503c"])

    ordered_monthly = monthly_summary.sort_index()
    first_month = ordered_monthly.index[0].strftime("%Y-%m")
    last_month = ordered_monthly.index[-1].strftime("%Y-%m")

    this_month_totals = get_window_totals(ordered_monthly, end_offset=0, months=1)
    prev_month_totals = get_window_totals(ordered_monthly, end_offset=1, months=1)

    rows = [
        {"\u6307\u6807": "\u7edf\u8ba1\u8303\u56f4", "\u503c": f"{first_month} \u81f3 {last_month}"},
        {"\u6307\u6807": "\u5e01\u79cd\u6570\u91cf", "\u503c": str(monthly_summary.shape[1])},
        {"\u6307\u6807": "\u672c\u6708Top3", "\u503c": format_top_symbols(get_window_totals(ordered_monthly, 0, 1), top_n=3)},
        {"\u6307\u6807": "\u4e0a\u4e2a\u6708Top3", "\u503c": format_top_symbols(get_window_totals(ordered_monthly, 1, 1), top_n=3)},
        {"\u6307\u6807": "\u4e0a\u4e09\u4e2a\u6708Top3", "\u503c": format_top_symbols(get_window_totals(ordered_monthly, 1, 3), top_n=3)},
        {"\u6307\u6807": "\u4e0a6\u4e2a\u6708Top3", "\u503c": format_top_symbols(get_window_totals(ordered_monthly, 1, 6), top_n=3)},
        {"\u6307\u6807": "\u4e0a12\u4e2a\u6708Top3", "\u503c": format_top_symbols(get_window_totals(ordered_monthly, 1, 12), top_n=3)},
        {"\u6307\u6807": "\u672c\u6708\u6b63\u8d39\u7387\u5e01\u79cd\u6570", "\u503c": str(int((this_month_totals > 0).sum()))},
        {"\u6307\u6807": "\u672c\u6708\u8d1f\u8d39\u7387\u5e01\u79cd\u6570", "\u503c": str(int((this_month_totals < 0).sum()))},
        {"\u6307\u6807": "\u672c\u6708\u96f6\u8d39\u7387\u5e01\u79cd\u6570", "\u503c": str(int((this_month_totals == 0).sum()))},
        {"\u6307\u6807": "\u4e0a\u6708\u6b63\u8d39\u7387\u5e01\u79cd\u6570", "\u503c": str(int((prev_month_totals > 0).sum()))},
        {"\u6307\u6807": "\u4e0a\u6708\u8d1f\u8d39\u7387\u5e01\u79cd\u6570", "\u503c": str(int((prev_month_totals < 0).sum()))},
        {"\u6307\u6807": "\u4e0a\u6708\u96f6\u8d39\u7387\u5e01\u79cd\u6570", "\u503c": str(int((prev_month_totals == 0).sum()))},
        {"\u6307\u6807": "\u4f4e\u6210\u4ea4\u91cf\u9608\u503c", "\u503c": f"{LOW_VOLUME_THRESHOLD:,.0f} USD"},
        {"\u6307\u6807": "\u4f4e\u6210\u4ea4\u91cf\u5e01\u79cd\u6570", "\u503c": str(len(low_volume_symbols))},
        {"\u6307\u6807": "\u4f4e\u6210\u4ea4\u91cf\u6807\u8bb0", "\u503c": f"\u5217\u540d\u5e26 {LOW_VOLUME_MARK}"},
    ]
    return pd.DataFrame(rows)


def fetch_volume_metrics(
    client: Client,
    symbols: List[str],
    contract_sizes: Dict[str, float],
    start_ts: int,
    group_timezone: str = GROUP_TIMEZONE,
) -> Tuple[pd.DataFrame, Dict[str, pd.DataFrame]]:
    now_ms = int(datetime.now(timezone.utc).timestamp() * 1000)
    window_span_ms = VOLUME_KLINE_MAX_WINDOW_DAYS * 86_400_000 - 1
    records: List[dict] = []
    volume_history: Dict[str, pd.DataFrame] = {}
    for symbol in symbols:
        klines: List[list] = []
        cursor = start_ts
        fetch_ok = True
        while cursor < now_ms:
            request_end = min(cursor + window_span_ms, now_ms)
            retries = 0
            batch = []
            while retries < MAX_RETRIES:
                try:
                    batch = client.futures_coin_klines(
                        symbol=symbol,
                        interval=Client.KLINE_INTERVAL_1DAY,
                        startTime=cursor,
                        endTime=request_end,
                        limit=VOLUME_KLINE_BATCH_LIMIT,
                    )
                    break
                except (BinanceAPIException, BinanceRequestException, RequestException) as exc:
                    retries += 1
                    wait = REQUEST_SLEEP_SECONDS * (2**retries)
                    print(f"[{symbol}] \u83b7\u53d6\u6210\u4ea4\u91cfK\u7ebf API\u9519\u8bef\uff0c{wait:.1f}s\u540e\u91cd\u8bd5 ({retries}/{MAX_RETRIES}): {exc}")
                    time.sleep(wait)
            else:
                print(f"[{symbol}] \u591a\u6b21\u5c1d\u8bd5\u83b7\u53d6\u6210\u4ea4\u91cf\u5931\u8d25\uff0c\u8df3\u8fc7\u8be5\u4ea4\u6613\u5bf9\u7684\u6210\u4ea4\u91cf\u8ba1\u7b97\u3002")
                fetch_ok = False
                klines = []
                break

            if not batch:
                next_cursor = request_end + 1
                if next_cursor <= cursor or request_end >= now_ms:
                    break
                cursor = next_cursor
                time.sleep(REQUEST_SLEEP_SECONDS)
                continue

            klines.extend(batch)
            next_cursor = request_end + 1
            if next_cursor <= cursor or request_end >= now_ms:
                break
            cursor = next_cursor

            time.sleep(REQUEST_SLEEP_SECONDS)

        if not fetch_ok or not klines:
            records.append({"Symbol": simplify_symbol(symbol), "AvgDailyUSDVolume": pd.NA, "FetchStatus": "Failed"})
            continue

        contract_size = contract_sizes.get(symbol, 1.0)
        daily_rows = []
        for entry in klines:
            # Binance COIN-M kline docs define:
            # entry[5] -> volume (contract count, "v")
            # entry[7] -> base asset volume ("q")
            # We persist contract_volume from v and derive usd_volume using contract_size.
            contract_volume = float(entry[5])
            usd_volume = contract_volume * contract_size
            daily_rows.append(
                {
                    "date": pd.to_datetime(int(entry[0]), unit="ms", utc=True).tz_convert(group_timezone).tz_localize(None),
                    "contract_volume": contract_volume,
                    "usd_volume": usd_volume,
                }
            )
        if daily_rows:
            volume_history[symbol] = pd.DataFrame(daily_rows).sort_values("date").drop_duplicates(subset="date", keep="last")

        recent_window = volume_history[symbol].tail(AVG_VOLUME_WINDOW_DAYS)
        avg_daily_usd = recent_window["usd_volume"].mean() if not recent_window.empty else pd.NA
        records.append({"Symbol": simplify_symbol(symbol), "AvgDailyUSDVolume": avg_daily_usd, "FetchStatus": "OK"})

    if not records:
        return pd.DataFrame(columns=["Symbol", "AvgDailyUSDVolume", "FetchStatus"]), volume_history

    df = pd.DataFrame(records)
    df["AvgDailyUSDVolume"] = pd.to_numeric(df["AvgDailyUSDVolume"], errors="coerce")
    df = (
        df.groupby("Symbol", as_index=False)
        .agg(
            AvgDailyUSDVolume=("AvgDailyUSDVolume", lambda s: s.sum(min_count=1)),
            FetchStatus=("FetchStatus", lambda s: "Failed" if "Failed" in set(s) else "OK"),
        )
    )
    df["StatusOrder"] = df["FetchStatus"].map({"OK": 0, "Failed": 1}).fillna(2)
    df.sort_values(["StatusOrder", "AvgDailyUSDVolume"], ascending=[True, False], na_position="last", inplace=True)
    df.drop(columns=["StatusOrder"], inplace=True)
    return df, volume_history


def resolve_lookback_days(lookback_days: int | None) -> int:
    if lookback_days is None:
        return FULL_LOOKBACK_YEARS * 365
    return max(int(lookback_days), 1)


def build_time_window(lookback_days: int) -> Tuple[datetime, int]:
    utc_now = datetime.now(timezone.utc)
    start_time = utc_now - timedelta(days=lookback_days)
    return start_time, int(start_time.timestamp() * 1000)


def compute_aggregate_refresh_bounds(daily_df: pd.DataFrame) -> Tuple[str, str]:
    if daily_df.empty:
        raise ValueError("daily_df must not be empty")

    min_date = pd.Timestamp(daily_df["date"].min()).normalize()
    max_date = pd.Timestamp(daily_df["date"].max()).normalize()
    week_start = min_date - pd.Timedelta(days=min_date.weekday())
    month_start = min_date.replace(day=1)
    agg_start = min(week_start, month_start)

    week_end = max_date + pd.Timedelta(days=(6 - max_date.weekday()))
    month_end = (max_date + pd.offsets.MonthEnd(0)).normalize()
    agg_end = max(week_end, month_end)
    return agg_start.strftime("%Y-%m-%d"), agg_end.strftime("%Y-%m-%d")


def refresh_symbol_aggregates(
    conn,
    symbol: str,
    changed_daily_df: pd.DataFrame,
    run_id: int,
) -> None:
    if changed_daily_df.empty:
        return

    agg_start, agg_end = compute_aggregate_refresh_bounds(changed_daily_df)
    aggregate_source = load_daily_funding_metrics(conn, symbol, start_date=agg_start, end_date=agg_end)
    persist_weekly_funding_metrics(conn, symbol, build_weekly_symbol_metrics(aggregate_source), run_id)
    persist_monthly_funding_metrics(conn, symbol, build_monthly_symbol_metrics(aggregate_source), run_id)


def auto_fit_columns(sheet, min_width: int = 9, max_width: int = 40, padding: int = 2) -> None:
    for col_idx in range(1, sheet.max_column + 1):
        max_len = 0
        for row_idx in range(1, sheet.max_row + 1):
            value = sheet.cell(row=row_idx, column=col_idx).value
            if value is None:
                continue
            text = value.strftime("%Y-%m-%d") if isinstance(value, datetime) else str(value)
            max_len = max(max_len, len(text))
        sheet.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + padding, min_width), max_width)


def apply_standard_layout(sheet, horizontal: str = "center") -> None:
    alignment = Alignment(horizontal=horizontal, vertical="center")
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            cell.alignment = alignment
    for col_idx in range(1, sheet.max_column + 1):
        header = sheet.cell(row=1, column=col_idx)
        font = copy(header.font)
        font.bold = True
        header.font = font


def apply_freeze_and_filter(sheet, freeze_cell: str, filter_last_row: int | None = None) -> None:
    sheet.freeze_panes = freeze_cell
    filter_last_row = sheet.max_row if filter_last_row is None else filter_last_row
    if sheet.max_column >= 1 and filter_last_row >= 1:
        sheet.auto_filter.ref = f"A1:{get_column_letter(sheet.max_column)}{filter_last_row}"


def add_monthly_heatmap(sheet, start_row: int, end_row: int, start_col: int, end_col: int) -> None:
    if end_row < start_row or end_col < start_col:
        return
    data_range = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"
    sheet.conditional_formatting.add(data_range, ColorScaleRule(start_type="min", start_color="FFF8D7DA", mid_type="percentile", mid_value=50, mid_color="FFFFFFFF", end_type="max", end_color="FFD4EDDA"))


def colorize_positive_negative(sheet, start_row: int, end_row: int, start_col: int, end_col: int) -> None:
    for row in sheet.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
        for cell in row:
            if not isinstance(cell.value, (int, float)):
                continue
            font = copy(cell.font)
            if cell.value > 0:
                font.color = POSITIVE_FONT.color
            elif cell.value < 0:
                font.color = NEGATIVE_FONT.color
            cell.font = font


def emphasize_monthly_top5(sheet, start_row: int, num_rows: int, num_cols: int) -> None:
    for row_idx in range(start_row, start_row + num_rows):
        values = []
        for col_idx in range(2, 2 + num_cols):
            cell = sheet.cell(row=row_idx, column=col_idx)
            if isinstance(cell.value, (int, float)):
                values.append((float(cell.value), col_idx))
        for _, col_idx in sorted(values, key=lambda item: item[0], reverse=True)[:5]:
            font = copy(sheet.cell(row=row_idx, column=col_idx).font)
            font.bold = True
            sheet.cell(row=row_idx, column=col_idx).font = font


def save_daily_excels(symbol_daily: Dict[str, pd.DataFrame]) -> None:
    OUTPUT_ROOT.mkdir(parents=True, exist_ok=True)
    temp_root = Path(tempfile.mkdtemp(prefix="daily_", dir=str(OUTPUT_ROOT)))
    try:
        for symbol, daily in symbol_daily.items():
            if daily.empty:
                continue

            output_path = temp_root / f"{symbol}_daily_funding.xlsx"
            final_output_path = DAILY_DIR / output_path.name
            daily_out = daily.copy()
            daily_out["date"] = pd.to_datetime(daily_out["date"]).dt.tz_localize(None)
            daily_out["funding_rate_7d_ma"] = daily_out["daily_funding_rate"].rolling(7, min_periods=1).mean()

            with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
                daily_out.to_excel(writer, index=False, sheet_name="DailyData")
                sheet = writer.sheets["DailyData"]
                max_row = sheet.max_row

                for row_idx in range(2, max_row + 1):
                    sheet.cell(row=row_idx, column=1).number_format = "yyyy-mm-dd"
                    sheet.cell(row=row_idx, column=2).number_format = "0.00%"
                    sheet.cell(row=row_idx, column=3).number_format = "0.00%"

                apply_freeze_and_filter(sheet, freeze_cell="A2")
                apply_standard_layout(sheet)
                auto_fit_columns(sheet, min_width=12, max_width=24)

                if max_row >= 2:
                    chart_start_row = max(max_row - DAILY_CHART_DAYS + 1, 2)
                    helper_start_col = 6
                    sheet.cell(row=1, column=helper_start_col, value="chart_date")
                    sheet.cell(row=1, column=helper_start_col + 1, value=sheet.cell(row=1, column=2).value)
                    sheet.cell(row=1, column=helper_start_col + 2, value=sheet.cell(row=1, column=3).value)
                    sheet.cell(row=1, column=helper_start_col + 3, value="zero_line")

                    helper_row = 2
                    recent_values: List[float] = []
                    for src_row in range(chart_start_row, max_row + 1):
                        date_value = sheet.cell(row=src_row, column=1).value
                        daily_value = sheet.cell(row=src_row, column=2).value
                        ma_value = sheet.cell(row=src_row, column=3).value
                        sheet.cell(row=helper_row, column=helper_start_col, value=date_value)
                        sheet.cell(row=helper_row, column=helper_start_col + 1, value=daily_value)
                        sheet.cell(row=helper_row, column=helper_start_col + 2, value=ma_value)
                        sheet.cell(row=helper_row, column=helper_start_col + 3, value=0.0)
                        for c in (helper_start_col + 1, helper_start_col + 2, helper_start_col + 3):
                            sheet.cell(row=helper_row, column=c).number_format = "0.00%"
                        if isinstance(daily_value, (int, float)):
                            recent_values.append(float(daily_value))
                        if isinstance(ma_value, (int, float)):
                            recent_values.append(float(ma_value))
                        helper_row += 1
                    helper_end_row = helper_row - 1

                    chart = LineChart()
                    chart.title = f"{simplify_symbol(symbol)} \u8fd130\u5929\u8d44\u91d1\u8d39\u7387\u4e0e7\u65e5\u5747\u7ebf"
                    chart.y_axis.title = "Funding Rate"
                    chart.x_axis.title = "Date"
                    chart.height = 8
                    chart.width = 18
                    chart.legend.position = "r"

                    data = Reference(sheet, min_col=helper_start_col + 1, max_col=helper_start_col + 3, min_row=1, max_row=helper_end_row)
                    categories = Reference(sheet, min_col=helper_start_col, min_row=2, max_row=helper_end_row)
                    chart.add_data(data, titles_from_data=True)
                    chart.set_categories(categories)
                    chart.x_axis.number_format = "yyyy-mm-dd"
                    chart.y_axis.number_format = "0.00%"
                    chart.x_axis.majorGridlines = None
                    chart.y_axis.majorGridlines = None
                    chart.plotVisOnly = False
                    if len(chart.series) >= 3:
                        zero_series = chart.series[2]
                        zero_series.graphicalProperties.line.solidFill = "000000"
                        zero_series.graphicalProperties.line.width = 25000

                    if recent_values:
                        min_v, max_v = min(recent_values), max(recent_values)
                        lower, upper = min(min_v, 0.0), max(max_v, 0.0)
                        span = upper - lower if upper - lower != 0 else 0.0005
                        pad = span * 0.1
                        chart.y_axis.scaling.min = lower - pad
                        chart.y_axis.scaling.max = upper + pad

                    sheet.add_chart(chart, "E2")
                    for col_idx in range(helper_start_col, helper_start_col + 4):
                        sheet.column_dimensions[get_column_letter(col_idx)].width = 12

            print(f"[{symbol}] \u65e5\u8d44\u91d1\u8d39\u7387\u5199\u5165: {final_output_path}")

        if DAILY_DIR.exists():
            shutil.rmtree(DAILY_DIR)
        temp_root.replace(DAILY_DIR)
        temp_root = DAILY_DIR
    finally:
        if temp_root.exists() and temp_root != DAILY_DIR:
            shutil.rmtree(temp_root, ignore_errors=True)


def add_symbol_trend_sheets(writer, monthly_summary: pd.DataFrame) -> Dict[str, str]:
    symbol_sheet_map: Dict[str, str] = {}
    used_names = set(writer.book.sheetnames)
    month_labels = [idx.strftime("%Y-%m") for idx in monthly_summary.index]

    for symbol in monthly_summary.columns:
        sheet_name = make_safe_sheet_name(f"Trend_{symbol}", used_names)
        symbol_sheet_map[symbol] = sheet_name
        ws = writer.book.create_sheet(title=sheet_name)
        ws.append(["Month", "FundingRatePct", "ZeroLine"])
        for m, value in zip(month_labels, monthly_summary[symbol].tolist()):
            ws.append([m, float(value) * 100.0, 0.0])

        for row_idx in range(2, ws.max_row + 1):
            ws.cell(row=row_idx, column=2).number_format = "0.0"
            ws.cell(row=row_idx, column=3).number_format = "0.0"

        chart = LineChart()
        chart.title = f"{symbol} \u6708\u5ea6\u8d44\u91d1\u8d39\u7387\u8d70\u52bf"
        chart.x_axis.title = "时间"
        chart.y_axis.title = "费率(%)"
        chart.height = 8
        chart.width = 16

        data = Reference(ws, min_col=2, max_col=3, min_row=1, max_row=ws.max_row)
        categories = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        chart.y_axis.number_format = "0.0"
        chart.y_axis.scaling.min = -3
        chart.y_axis.scaling.max = 3
        chart.y_axis.majorUnit = 0.5
        chart.x_axis.majorGridlines = None
        chart.y_axis.majorGridlines = None
        if len(chart.series) >= 1:
            chart.series[0].graphicalProperties.line.solidFill = "1F4E79"
            chart.series[0].graphicalProperties.line.width = 25000
        if len(chart.series) >= 2:
            chart.series[1].graphicalProperties.line.solidFill = "000000"
            chart.series[1].graphicalProperties.line.width = 20000
        ws.add_chart(chart, "D2")

        apply_freeze_and_filter(ws, freeze_cell="A2")
        apply_standard_layout(ws)
        auto_fit_columns(ws, min_width=10, max_width=20)

    return symbol_sheet_map


def save_monthly_summary(monthly_summary: pd.DataFrame, stats_table: pd.DataFrame, label: str = "", avg_volume_table: pd.DataFrame | None = None) -> Path:
    OUTPUT_ROOT.mkdir(parents=True, exist_ok=True)
    timestamp_str = datetime.now().strftime(SUMMARY_TIME_FORMAT)
    label_part = f"{label}_" if label else ""
    path = OUTPUT_ROOT / f"{SUMMARY_BASENAME}_{label_part}{timestamp_str}.xlsx"

    low_volume_symbols: Set[str] = set()
    if avg_volume_table is not None and not avg_volume_table.empty:
        low_volume_symbols = set(
            avg_volume_table.loc[
                (avg_volume_table["FetchStatus"] == "OK") & (avg_volume_table["AvgDailyUSDVolume"] < LOW_VOLUME_THRESHOLD),
                "Symbol",
            ].astype(str)
        )

    summary_with_total = monthly_summary.copy()
    summary_with_total.loc["\u603b\u8ba1"] = summary_with_total.sum()
    display_df = summary_with_total.copy()
    display_df.index = [idx if isinstance(idx, str) else idx.strftime("%Y-%m") for idx in display_df.index]
    display_df.columns = [f"{col}{LOW_VOLUME_MARK}" if col in low_volume_symbols else col for col in display_df.columns]

    ranking_df = stats_table.copy()
    ranking_df.index.name = "\u533a\u95f4"
    overview_df = build_overview_table(monthly_summary, low_volume_symbols)

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        overview_df.to_excel(writer, sheet_name=OVERVIEW_SHEET_NAME, index=False)
        display_df.to_excel(writer, sheet_name=MONTHLY_SHEET_NAME)
        ranking_df.to_excel(writer, sheet_name=RANKING_SHEET_NAME)

        if avg_volume_table is not None:
            volume_df = avg_volume_table.copy() if not avg_volume_table.empty else pd.DataFrame(columns=["Symbol", "AvgDailyUSDVolume", "FetchStatus"])
            volume_df["LowVolume"] = volume_df.apply(
                lambda row: "Yes"
                if row.get("FetchStatus") == "OK" and pd.notna(row.get("AvgDailyUSDVolume")) and float(row["AvgDailyUSDVolume"]) < LOW_VOLUME_THRESHOLD
                else "",
                axis=1,
            )
            volume_df.to_excel(writer, sheet_name=VOLUME_SHEET_NAME, index=False)

        monthly_sheet = writer.sheets[MONTHLY_SHEET_NAME]
        max_row = monthly_sheet.max_row
        max_col = monthly_sheet.max_column
        month_rows = len(monthly_summary)

        for row in monthly_sheet.iter_rows(min_row=2, max_row=max_row, min_col=2, max_col=max_col):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "0.00%"

        for col_idx in range(2, max_col + 1):
            header_cell = monthly_sheet.cell(row=1, column=col_idx)
            if isinstance(header_cell.value, str) and header_cell.value.endswith(LOW_VOLUME_MARK):
                header_cell.fill = LOW_VOLUME_FILL

        add_monthly_heatmap(monthly_sheet, start_row=2, end_row=month_rows + 1, start_col=2, end_col=max_col)
        colorize_positive_negative(monthly_sheet, start_row=2, end_row=max_row, start_col=2, end_col=max_col)
        emphasize_monthly_top5(monthly_sheet, start_row=2, num_rows=month_rows, num_cols=monthly_summary.shape[1])
        apply_freeze_and_filter(monthly_sheet, freeze_cell="B2")
        apply_standard_layout(monthly_sheet)
        auto_fit_columns(monthly_sheet, min_width=10, max_width=24)

        symbol_sheet_map = add_symbol_trend_sheets(writer, monthly_summary)
        for col_idx, symbol in enumerate(monthly_summary.columns, start=2):
            header_cell = monthly_sheet.cell(row=1, column=col_idx)
            target_sheet = symbol_sheet_map.get(symbol)
            if target_sheet:
                header_cell.hyperlink = f"#{target_sheet}!A1"
                header_cell.style = "Hyperlink"

        overview_sheet = writer.sheets[OVERVIEW_SHEET_NAME]
        apply_freeze_and_filter(overview_sheet, freeze_cell="A2")
        apply_standard_layout(overview_sheet, horizontal="left")
        for row_idx in range(2, overview_sheet.max_row + 1):
            overview_sheet.cell(row=row_idx, column=1).alignment = Alignment(horizontal="center", vertical="center")
        auto_fit_columns(overview_sheet, min_width=12, max_width=80)

        ranking_sheet = writer.sheets[RANKING_SHEET_NAME]
        apply_freeze_and_filter(ranking_sheet, freeze_cell="B2")
        apply_standard_layout(ranking_sheet)
        auto_fit_columns(ranking_sheet, min_width=10, max_width=40)

        if avg_volume_table is not None:
            volume_sheet = writer.sheets[VOLUME_SHEET_NAME]
            for row_idx in range(2, volume_sheet.max_row + 1):
                volume_value = volume_sheet.cell(row=row_idx, column=2).value
                volume_sheet.cell(row=row_idx, column=2).number_format = "#,##0"
                try:
                    numeric_value = float(volume_value)
                except (TypeError, ValueError):
                    continue
                status_value = volume_sheet.cell(row=row_idx, column=3).value
                if status_value == "OK" and numeric_value < LOW_VOLUME_THRESHOLD:
                    for col_idx in range(1, volume_sheet.max_column + 1):
                        volume_sheet.cell(row=row_idx, column=col_idx).fill = LOW_VOLUME_FILL
            apply_freeze_and_filter(volume_sheet, freeze_cell="A2")
            apply_standard_layout(volume_sheet)
            auto_fit_columns(volume_sheet, min_width=12, max_width=28)

    print(f"\u6708\u5ea6\u7edf\u8ba1\u5199\u5165: {path}")
    return path


def main() -> None:
    parser = argparse.ArgumentParser(description="Collect Binance COIN-M funding and volume into Excel + SQLite.")
    parser.add_argument("--timezone", default=GROUP_TIMEZONE, help="Aggregation timezone, e.g. Asia/Shanghai or UTC.")
    parser.add_argument(
        "--lookback-days",
        type=int,
        default=None,
        help="Funding incremental window in days. Omit to keep the old 3-year full refresh behavior.",
    )
    parser.add_argument(
        "--volume-lookback-days",
        type=int,
        default=None,
        help="Volume incremental window in days. Defaults to max(lookback_days, 30 + buffer).",
    )
    args = parser.parse_args()

    client = Client()
    funding_lookback_days = resolve_lookback_days(args.lookback_days)
    volume_lookback_days = resolve_lookback_days(args.volume_lookback_days)
    volume_lookback_days = max(volume_lookback_days, funding_lookback_days, AVG_VOLUME_WINDOW_DAYS + DEFAULT_VOLUME_LOOKBACK_EXTRA_DAYS)
    utc_now = datetime.now(timezone.utc)
    funding_start_time, funding_start_ts = build_time_window(funding_lookback_days)
    volume_start_time, volume_start_ts = build_time_window(volume_lookback_days)
    print(
        f"\u5f00\u59cb\u6293\u53d6\u5e01\u672c\u4f4d\u5408\u7ea6\u8d44\u91d1\u8d39\u7387\uff0c"
        f"funding\u7a97\u53e3: {funding_start_time.date()} - {utc_now.date()} ; "
        f"volume\u7a97\u53e3: {volume_start_time.date()} - {utc_now.date()}"
    )

    symbols, contract_sizes = get_coin_perpetual_symbols(client)
    if not symbols:
        print("\u672a\u83b7\u53d6\u5230\u4efb\u4f55\u5e01\u672c\u4f4d\u6c38\u7eed\u5408\u7ea6\u4ea4\u6613\u5bf9\u3002")
        return
    print(f"\u5171 {len(symbols)} \u4e2a\u4ea4\u6613\u5bf9\u3002")

    run_id = None
    symbol_daily: Dict[str, pd.DataFrame] = {}
    skipped_symbols: List[str] = []
    quality_audits: List[Dict[str, object]] = []
    try:
        with sqlite_connection() as conn:
            initialize_database(conn)
            run_id = create_collector_run(
                conn,
                lookback_years=max(1, math.ceil(funding_lookback_days / 365)),
                symbol_count=len(symbols),
            )
            upsert_symbols(conn, symbols, contract_sizes)

            for idx, symbol in enumerate(symbols, start=1):
                print(f"[{idx}/{len(symbols)}] \u6293\u53d6 {symbol} ...")
                try:
                    funding_df = fetch_symbol_funding_history(client, symbol, start_ts=funding_start_ts)
                except DataFetchIncompleteError as exc:
                    print(str(exc))
                    skipped_symbols.append(symbol)
                    continue

                symbol_daily[symbol] = compute_daily_funding(funding_df, group_timezone=args.timezone)
                quality_audits.append(
                    build_funding_quality_audit(
                        symbol,
                        funding_df,
                        symbol_daily[symbol],
                        raw_row_count=int(funding_df.attrs.get("raw_row_count", len(funding_df))),
                    )
                )

                persist_raw_funding_rates(conn, symbol, funding_df, run_id)
                persist_daily_funding_metrics(conn, symbol, symbol_daily[symbol], run_id)
                refresh_symbol_aggregates(conn, symbol, symbol_daily[symbol], run_id)
                persist_funding_quality_audit(conn, run_id, quality_audits[-1])

            if not symbol_daily:
                finalize_collector_run(
                    conn,
                    run_id=run_id,
                    status="failed",
                    skipped_symbol_count=len(skipped_symbols),
                    notes="all funding fetches failed",
                )
                print("\u6240\u6709\u4ea4\u6613\u5bf9\u7684\u8d44\u91d1\u8d39\u7387\u6293\u53d6\u5747\u5931\u8d25\uff0c\u672a\u751f\u6210\u4efb\u4f55\u8f93\u51fa\u3002")
                return

            if skipped_symbols:
                print(f"\u5df2\u8df3\u8fc7 {len(skipped_symbols)} \u4e2a\u6570\u636e\u4e0d\u5b8c\u6574\u7684\u4ea4\u6613\u5bf9: {', '.join(skipped_symbols)}")
            warning_symbols = [str(entry["symbol"]) for entry in quality_audits if str(entry["status"]) != "ok"]

            full_symbol_daily = {
                symbol: load_daily_funding_metrics(conn, symbol)
                for symbol in symbol_daily.keys()
            }
            save_daily_excels(full_symbol_daily)
            avg_volume_table, volume_history = fetch_volume_metrics(
                client,
                list(full_symbol_daily.keys()),
                contract_sizes,
                start_ts=volume_start_ts,
                group_timezone=args.timezone,
            )
            for symbol, volume_df in volume_history.items():
                persist_daily_volume_metrics(conn, symbol, volume_df, run_id)
                persist_volume_quality_audit(
                    conn,
                    run_id,
                    build_volume_quality_audit(symbol, volume_df, expected_window_days=volume_lookback_days),
                )
            for row in avg_volume_table.itertuples(index=False):
                if str(row.FetchStatus) != "OK":
                    failed_symbol = f"{str(row.Symbol)}USD_PERP"
                    persist_volume_quality_audit(conn, run_id, build_failed_volume_quality_audit(failed_symbol, "volume fetch failed or no kline rows"))

            monthly_37 = compute_monthly_summary(full_symbol_daily, periods=37, group_timezone=args.timezone)
            if monthly_37.empty:
                finalize_collector_run(
                    conn,
                    run_id=run_id,
                    status="failed",
                    skipped_symbol_count=len(skipped_symbols),
                    notes="monthly summary empty",
                )
                print("\u672a\u751f\u6210\u6708\u5ea6\u7edf\u8ba1\u6570\u636e\u3002")
                return

            high_liquidity_symbols = [
                symbol
                for symbol, daily_volume in volume_history.items()
                if not daily_volume.empty and daily_volume["usd_volume"].mean() >= 150_000_000
            ]
            persist_market_snapshots(conn, monthly_37, FOCUS_BASKET, high_liquidity_symbols, run_id)

            save_monthly_summary(monthly_37, build_recent_top_table(monthly_37), label="\u8fd137\u4e2a\u6708", avg_volume_table=avg_volume_table)

            monthly_24 = monthly_37.tail(24)
            if not monthly_24.empty:
                monthly_24 = monthly_24[monthly_24.sum().sort_values(ascending=False).index]
            save_monthly_summary(monthly_24, build_recent_top_table(monthly_24), label="\u8fd124\u4e2a\u6708", avg_volume_table=avg_volume_table)

            recent_12 = monthly_37.iloc[-13:-1]
            if not recent_12.empty:
                recent_12 = recent_12[recent_12.sum().sort_values(ascending=False).index]
            save_monthly_summary(recent_12, build_recent_top_table(recent_12), label="\u8fd112\u6708", avg_volume_table=avg_volume_table)

            finalize_collector_run(
                conn,
                run_id=run_id,
                status="completed",
                skipped_symbol_count=len(skipped_symbols),
                notes=(
                    f"persisted {len(symbol_daily)} symbols; audit warnings={len(warning_symbols)}; "
                    f"timezone={args.timezone}; funding_lookback_days={funding_lookback_days}; "
                    f"volume_lookback_days={volume_lookback_days}"
                ),
            )
            print("\u5df2\u5b8c\u6210 SQLite \u5199\u5165\u4e0e Excel \u5bfc\u51fa\u3002")
    except Exception as exc:
        if run_id is not None:
            with sqlite_connection() as conn:
                initialize_database(conn)
                finalize_collector_run(
                    conn,
                    run_id=run_id,
                    status="failed",
                    skipped_symbol_count=len(skipped_symbols),
                    notes=str(exc)[:500],
                )
        raise


if __name__ == "__main__":
    main()
